# Import required libraries
import snowflake.connector
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import numpy as np
from io import BytesIO
from openpyxl import Workbook
import datetime


# Displaying images on the front end
from PIL import Image
image = Image.open('images/DeltaPacific_Logo.jpg')
st.image(image, caption='',width=200)

st.title("Delta Pacific Beverage")
st.header("Gap Report and Analysis")



def format_sales_report(workbook):
    # Delete all sheets except SALES REPORT
    for sheet_name in workbook.sheetnames:
        if sheet_name != 'SALES REPORT':
            workbook.remove(workbook[sheet_name])

    # Select the SALES REPORT sheet
    ws = workbook['SALES REPORT']

    # Delete row 2
    ws.delete_rows(2)

    # Delete column H
    ws.delete_cols(8)

    # Remove all hyphens from column F
    for cell in ws['F']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('-', '')

    # Create a new column for store name
    ws.insert_cols(2)
    ws.cell(row=1, column=2, value='STORE NAME')

    # Copy values before the # to store name column
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if '#' in str(cell.value):
                cell_offset = ws.cell(row=cell.row, column=2)
                store_name = str(cell.value).split('#')[0].replace("'", "")
                cell_offset.value = store_name

    # Remove column C
    ws.delete_cols(3)

    # Replace all commas with spaces in column B
    for cell in ws['B']:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace(',', ' ')

    # Remove all 's in column B
    for cell in ws['B']:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace(" 's", "")

    # Replace all commas with spaces in column E
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace(',', ' ')

   
    # Remove all commas from column C
    for cell in ws['C']:
        if cell.value is not None:
            cell.value = str(cell.value).replace(',', ' ')
            
            
            
    # Remove all Is Null from column F
    for cell in ws['F']:
        if cell.value is not None:
            cell.value = str(cell.value).replace('Is Null', '0')

    # Format column G as number with no decimals
    for cell in ws['G'][1:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = numbers.FORMAT_NUMBER
        elif isinstance(cell.value, str):
            cell.number_format = numbers.FORMAT_NUMBER
            try:
                cell.value = float(cell.value.replace(",", ""))
            except ValueError:
                pass
                
    
    

    return workbook


# Upload the workbook
uploaded_file = st.file_uploader("Upload freshly ran sales report from Encompass", type=["xlsx", "xls"])

if uploaded_file is not None:
    # Load the workbook
    workbook = openpyxl.load_workbook(uploaded_file)

    # Show the Reformat button
    if st.button("Reformat"):
        # Format the sales report
        new_workbook = format_sales_report(workbook)

        # Download the formatted file
        new_filename = 'formatted_' + uploaded_file.name
        stream = BytesIO()
        new_workbook.save(stream)
        stream.seek(0)
        st.download_button(label="Download formatted file", data=stream.read(), file_name=new_filename, mime='application/vnd.ms-excel')
    



def write_to_snowflake(df, warehouse, database, schema, env):
    
    
    # read Excel file into pandas DataFrame
    df = pd.read_excel(uploaded_file)
    
   
    # replace NaN values with "NULL"
    df.fillna(value=np.nan, inplace=True)
    



  
    
    # Load Snowflake credentials from the secrets.toml file
    snowflake_creds = st.secrets["snowflake"]

    # Establish a new connection to Snowflake
    conn = snowflake.connector.connect(
    account=snowflake_creds["account"],
    user=snowflake_creds["user"],
    password=snowflake_creds["password"],
    warehouse=snowflake_creds["warehouse"],
    database=snowflake_creds["database"],
    schema=snowflake_creds["schema"]
    )
    
    
    if env == "production":
        table_name = "SALES_REPORT"
        if not st.sidebar.checkbox("Are you sure you want to import data into the production environment?"):
            st.warning("Data import has been cancelled")
            return
    elif env == "testing":
        table_name = "TMP_TABLE"
    else:
        st.error("Invalid environment selected")
    return


    
 
    # write DataFrame to Snowflake
    cursor = conn.cursor()
    sql_query = "CREATE OR REPLACE TABLE SALES_REPORT AS SELECT \
    CAST(STORE_NUMBER AS NUMBER) AS STORE_NUMBER, \
    CAST(STORE_NAME AS VARCHAR) AS STORE_NAME, \
    CAST(ADDRESS AS VARCHAR) AS ADDRESS, \
    CAST(SALESPERSON AS VARCHAR) AS SALESPERSON, \
    CAST(PRODUCT_NAME AS VARCHAR) AS PRODUCT_NAME, \
    CAST(UPC AS VARCHAR) AS UPC, \
    CAST(PURCHASED_YES_NO AS VARCHAR) AS PURCHASED_YES_NO \
    FROM (VALUES {}) \
    AS tmp(STORE_NUMBER, STORE_NAME, ADDRESS, SALESPERSON, PRODUCT_NAME, UPC, PURCHASED_YES_NO);".format(
        ', '.join([str(tuple(df.iloc[i].fillna(np.nan).values)) for i in range(len(df))])
    )
    #st.write(sql_query)  # print the SQL query
    cursor.execute(sql_query)
    cursor.close()
    conn.close()
    st.write("Data has been imported into Snowflake!")


# create file uploader
uploaded_file = st.file_uploader("UPLOAD CURRENT SALES REPORT AFTER IT HAS BEEN FORMATED", type=["xlsx"])
env = st.sidebar.selectbox("Select environment:", ["production", "testing"])

# check if file was uploaded
if uploaded_file:
    # read Excel file into pandas DataFrame
    df = pd.read_excel(uploaded_file)
    print(df.columns)
    # display DataFrame in Streamlit
    st.dataframe(df)

    # write DataFrame to Snowflake on button click
    if st.button("Import into Snowflake"):
        with st.spinner('Uploading data to Snowflake ...'):
            write_to_snowflake(df, "COMPUTE_WH", "datasets", "DATASETS", env)

#import streamlit as st
#import pandas as pd
#import base64
#import snowflake.connector
#from io import BytesIO


def get_table_download_link(df):
    """
    Generates a link allowing the data in a given pandas dataframe to be downloaded in CSV format.
    """
    """
    csv = df.to_csv(index=False)
    b64 = base64.b64encode(csv.encode()).decode()
    href = f'<a href="data:file/csv;base64,{b64}" download="gap_report.csv">Click here to download the Gap Report!</a>'
    return href
    """
    env = st.sidebar.selectbox("Select environment:", ["production", "testing"])

def create_gap_report(conn):
    """
    Retrieves data from a Snowflake view and creates a button to download the data as a CSV report.
    """
    # Execute SQL query and retrieve data from the Gap_Report view
    query = "SELECT * FROM Gap_Report"
    df = pd.read_sql(query, conn)

    # Write the updated dataframe to a temporary file
    temp_file_path = 'temp.xlsx'
    df.to_excel(temp_file_path, index=False)
    
    # Add a download button
    with open(temp_file_path, 'rb') as f:
        bytes_data = f.read()
        today = datetime.datetime.today().strftime('%Y-%m-%d') # get current date in YYYY-MM-DD format
        file_name = f"Gap_Report_{today}.xlsx" # insert current date into file name

        st.download_button(label="Download Gap Report", data=bytes_data, file_name=file_name, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        
    st.write("File will Be Downloaded to Your Local Download Folder")
    
    
    st.dataframe(df)
   

# Load Snowflake credentials from the secrets.toml file
snowflake_creds = st.secrets["snowflake"]
with st.sidebar:
    # Establish a new connection to Snowflake
    conn = snowflake.connector.connect(
        account=snowflake_creds["account"],
        user=snowflake_creds["user"],
        password=snowflake_creds["password"],
        warehouse=snowflake_creds["warehouse"],
        database=snowflake_creds["database"],
        schema=snowflake_creds["schema"]
    )

    if st.sidebar.button('Generate Gap Report :beers:'):
        with st.spinner('Generating report...'):
            create_gap_report(conn)
